"""Build an evidence-preserving RD testing integration from a source workbook.

Operational source and machine outputs use ASCII keys and English labels.
Human-facing report text is written separately in Spanish.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\issvk\Downloads\Testeo 2025.xlsx")
OUT_DIR = Path(r"C:\Users\issvk\claude_sesiones_recuperadas")
STAMP = "2026-08-12"
SOURCE_COPY = OUT_DIR / "Testeo 2025.source.xlsx"
JSON_OUT = OUT_DIR / f"rd_testeos_eventos_2025_evidence_{STAMP}.json"
XLSX_OUT = OUT_DIR / f"rd_testeos_eventos_2025_integrated_{STAMP}.xlsx"
REPORT_OUT = OUT_DIR / f"rd_testeos_eventos_2025_informe_{STAMP}.md"


RAW_COLUMNS = [
    "substance_raw",
    "format_raw",
    "test_1_raw",
    "result_1_raw",
    "test_2_raw",
    "result_2_raw",
    "test_3_raw",
    "result_3_raw",
    "test_4_raw",
    "result_4_raw",
    "extra_1_raw",
]

SUBSTANCE_MAP = {
    "mdma": ("mdma", "canonical_existing_registry"),
    "extasis": ("mdma", "candidate_alias"),
    "ketamina": ("ketamine", "canonical_existing_registry"),
    "ketamina+m": ("ketamine", "mixture_or_typo_candidate"),
    "cocaina": ("cocaine", "candidate_alias"),
    "2c-b": ("two_c_b", "canonical_existing_registry"),
    "tusi": ("tusi", "canonical_existing_registry"),
    "tussi": ("tusi", "candidate_alias"),
    "tus si": ("tusi", "candidate_alias"),
    "tu si": ("tusi", "candidate_alias"),
    "ghb": ("ghb_gbl", "candidate_alias"),
    "cannabis": ("cannabis", "canonical_existing_registry"),
    "mefedrona": ("mephedrone", "canonical_existing_registry"),
    "desconocido": ("unknown", "explicit_unknown"),
}

REAGENT_MAP = {
    "marquis": ("marquis", "canonical_registry"),
    "ehrlich": ("ehrlich", "canonical_registry"),
    "liebermann": ("liebermann", "canonical_registry"),
    "lieberman": ("liebermann", "candidate_alias"),
    "morris": ("morris", "canonical_registry"),
    "froehde": ("froehde", "canonical_registry"),
    "simón": ("simons", "candidate_alias"),
    "simon": ("simons", "candidate_alias"),
    "simons": ("simons", "candidate_alias"),
    "simon's": ("simons", "candidate_alias"),
    "simon s": ("simons", "candidate_alias"),
    "simons azul": ("simons", "alias_with_embedded_result_word"),
    "zimmermann": ("zimmermann", "canonical_registry"),
    "zimmerman": ("zimmermann", "candidate_alias"),
    "zimermann": ("zimmermann", "candidate_alias"),
    "mandelin": ("mandelin", "canonical_registry"),
    "robadope": ("robadope", "canonical_registry"),
    "robandole": ("robadope", "candidate_alias"),
    "robándole": ("robadope", "candidate_alias"),
    "maquis": ("marquis", "candidate_alias"),
}

RESULT_MAP = {
    "sin reaccion": "no_reaction_candidate",
    "ninguna": "no_reaction_candidate",
    "niguna": "no_reaction_candidate",
    "no reaction": "no_reaction_candidate",
}


def fold(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.strip().lower())


def text_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return f"{prefix}-{hashlib.sha256(raw.encode('utf-8')).hexdigest()[:16]}"


def parse_event_date(sheet_name: str) -> dict[str, Any]:
    """Parse only unambiguous dates; preserve uncertain tokens as evidence."""
    name = sheet_name.strip()
    patterns = [
        (r"(?<!\d)(\d{2})[-/.](\d{2})[-/.](\d{4})(?!\d)", "dd_mm_yyyy"),
        (r"(?<!\d)(\d{2})[-/.](\d{2})[-/.](\d{2})(?!\d)", "dd_mm_yy"),
        (r"(?<!\d)(\d{2})(\d{2})(\d{4})(?!\d)", "ddmmyyyy"),
        (r"(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)", "ddmmyy"),
    ]
    for pattern, style in patterns:
        match = re.search(pattern, name)
        if not match:
            continue
        day, month, year = (int(x) for x in match.groups())
        year += 2000 if year < 100 else 0
        try:
            parsed = date(year, month, day)
        except ValueError:
            continue
        return {
            "date_raw_token": match.group(0),
            "date_iso_candidate": parsed.isoformat(),
            "date_status": "parsed_candidate",
            "date_parse_style": style,
            "date_confidence": "medium" if style in {"dd_mm_yyyy", "dd_mm_yy", "ddmmyyyy", "ddmmyy"} else "low",
        }
    compact = re.search(r"(?<!\d)(\d{4})(?!\d)", name)
    if compact:
        token = compact.group(1)
        day, month = int(token[:2]), int(token[2:])
        if 1 <= day <= 31 and 1 <= month <= 12:
            return {
                "date_raw_token": token,
                "date_iso_candidate": None,
                "date_status": "partial_day_month_compact",
                "date_parse_style": "ddmm_without_year",
                "date_confidence": "low",
            }
        if 1 <= int(token[0]) <= 9 and 1 <= int(token[1]) <= 9:
            return {
                "date_raw_token": token,
                "date_iso_candidate": None,
                "date_status": "ambiguous_compact_numeric_token",
                "date_parse_style": "unresolved_compact",
                "date_confidence": "none",
            }
    partial = re.search(r"(?<!\d)(\d{1,2})[-/.](\d{1,2})(?!\d)", name)
    if partial:
        return {
            "date_raw_token": partial.group(0),
            "date_iso_candidate": None,
            "date_status": "partial_day_month_only",
            "date_parse_style": "dd_mm_partial",
            "date_confidence": "low",
        }
    return {
        "date_raw_token": None,
        "date_iso_candidate": None,
        "date_status": "not_found",
        "date_parse_style": None,
        "date_confidence": "none",
    }


def normalize_substance(raw: str | None) -> tuple[str | None, str]:
    key = fold(raw)
    if not key:
        return None, "missing"
    if key in SUBSTANCE_MAP:
        return SUBSTANCE_MAP[key]
    if key in {"sustancia"}:
        return None, "repeated_header"
    if key in {"polvo blanco", "lamborghini dorada", "freedom explicito"}:
        return "unknown", "misplaced_or_unresolved_candidate"
    return "unknown", "unresolved_candidate"


def normalize_reagent(raw: str | None) -> tuple[str | None, str]:
    key = fold(raw)
    if not key:
        return None, "missing"
    if key in {"test 1", "test 2", "test 3", "test 4"}:
        return None, "repeated_header"
    if key in REAGENT_MAP:
        return REAGENT_MAP[key]
    if key == "\ufffc":
        return None, "object_placeholder"
    return "unknown", "unresolved_candidate"


def normalize_result(raw: str | None) -> tuple[str | None, str]:
    key = fold(raw)
    if not key:
        return None, "missing"
    return RESULT_MAP.get(key), "observed_wording_only" if key not in RESULT_MAP else "no_reaction_wording_candidate"


def make_event_record(ws_title: str, sheet_index: int, group: dict[str, Any], source_period: str) -> dict[str, Any]:
    parsed = parse_event_date(ws_title)
    label = re.sub(r"^(copy of\s+)+", "", ws_title.strip(), flags=re.IGNORECASE)
    event_id = stable_id("event", sheet_index, ws_title)
    outside = bool(parsed["date_iso_candidate"] and not parsed["date_iso_candidate"].startswith(source_period))
    return {
        "event_id": event_id,
        "source_sheet_index": sheet_index,
        "source_sheet_name": ws_title,
        "event_label_candidate": label,
        "event_label_status": "source_sheet_name_only",
        "source_period_label": source_period,
        **parsed,
        "outside_filename_period_candidate": outside,
        "is_source_copy_candidate": bool(re.match(r"^copy of\s+", ws_title.strip(), re.I)),
        "duplicate_group_id": group["duplicate_group_id"],
        "duplicate_group_size": group["size"],
        "duplicate_status": group["status"],
        "duplicate_canonical_sheet_candidate": group["canonical_sheet"],
        "venue_id": None,
        "venue_name": None,
        "producer_id": None,
        "producer_name": None,
        "link_status": "unlinked",
        "link_evidence_ref": None,
        "link_confidence": "none",
        "link_review_status": "pending_human_link",
    }


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    source_bytes = SOURCE.read_bytes()
    source_hash = hashlib.sha256(source_bytes).hexdigest()
    if not SOURCE_COPY.exists():
        shutil.copy2(SOURCE, SOURCE_COPY)

    wb = load_workbook(SOURCE, read_only=True, data_only=False)
    source_period = "2025"
    sheet_data: dict[str, list[tuple[int, list[str | None]]]] = {}
    sheet_hash: dict[str, str] = {}
    for ws in wb.worksheets:
        rows: list[tuple[int, list[str | None]]] = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [text_value(v) for v in list(row[:11])]
            if not any(v is not None and v.strip() for v in values):
                continue
            rows.append((row_index, values))
        sheet_data[ws.title] = rows
        payload = json.dumps([values for _, values in rows], ensure_ascii=True, sort_keys=True)
        sheet_hash[ws.title] = hashlib.sha256(payload.encode("utf-8")).hexdigest()

    groups: dict[str, list[str]] = defaultdict(list)
    for name, digest in sheet_hash.items():
        groups[digest].append(name)
    group_by_sheet: dict[str, dict[str, Any]] = {}
    for digest, names in groups.items():
        group_id = f"dup-{digest[:16]}"
        status = "exact_sheet_duplicate_candidate" if len(names) > 1 else "unique_source_sheet"
        for name in names:
            group_by_sheet[name] = {
                "duplicate_group_id": group_id,
                "size": len(names),
                "status": status,
                "canonical_sheet": names[0],
            }

    events: list[dict[str, Any]] = []
    event_by_sheet: dict[str, dict[str, Any]] = {}
    for index, ws in enumerate(wb.worksheets, start=1):
        event = make_event_record(ws.title, index, group_by_sheet[ws.title], source_period)
        events.append(event)
        event_by_sheet[ws.title] = event

    test_rows: list[dict[str, Any]] = []
    observations: list[dict[str, Any]] = []
    substance_counts: Counter[str] = Counter()
    reagent_counts: Counter[str] = Counter()
    result_counts: Counter[str] = Counter()
    row_status_counts: Counter[str] = Counter()
    map_substance_raw: dict[str, dict[str, Any]] = {}
    map_reagent_raw: dict[str, dict[str, Any]] = {}

    for event in events:
        for source_row, values in sheet_data[event["source_sheet_name"]]:
            raw = dict(zip(RAW_COLUMNS, values + [None] * (len(RAW_COLUMNS) - len(values))))
            substance_id, substance_status = normalize_substance(raw["substance_raw"])
            if substance_status == "repeated_header":
                row_status = "repeated_header"
            elif not any(v is not None and v.strip() for v in values):
                row_status = "empty"
            elif substance_status in {"missing", "unresolved_candidate", "misplaced_or_unresolved_candidate"}:
                row_status = "data_with_unresolved_substance"
            else:
                row_status = "data"
            row_status_counts[row_status] += 1
            if raw["substance_raw"]:
                key = raw["substance_raw"]
                substance_counts[key] += 1
                map_substance_raw.setdefault(key, {"raw_label": key, "count": 0, "normalized_id": substance_id, "mapping_status": substance_status})
                map_substance_raw[key]["count"] += 1
            test_id = stable_id("test", source_hash, event["event_id"], source_row)
            test = {
                "test_id": test_id,
                "event_id": event["event_id"],
                "source_sheet_name": event["source_sheet_name"],
                "source_row": source_row,
                "row_status": row_status,
                "substance_raw": raw["substance_raw"],
                "substance_normalized_candidate": substance_id,
                "substance_map_status": substance_status,
                "format_raw": raw["format_raw"],
                "test_1_raw": raw["test_1_raw"],
                "result_1_raw": raw["result_1_raw"],
                "test_2_raw": raw["test_2_raw"],
                "result_2_raw": raw["result_2_raw"],
                "test_3_raw": raw["test_3_raw"],
                "result_3_raw": raw["result_3_raw"],
                "test_4_raw": raw["test_4_raw"],
                "result_4_raw": raw["result_4_raw"],
                "extra_1_raw": raw["extra_1_raw"],
                "source_duplicate_group_id": event["duplicate_group_id"],
                "source_duplicate_status": event["duplicate_status"],
                "interpretation_policy": "observed_color_only_not_identity_purity_or_dose",
            }
            test_rows.append(test)
            for ordinal in range(1, 5):
                reagent_raw = raw[f"test_{ordinal}_raw"]
                result_raw = raw[f"result_{ordinal}_raw"]
                if reagent_raw is None and result_raw is None:
                    continue
                reagent_id, reagent_status = normalize_reagent(reagent_raw)
                result_candidate, result_status = normalize_result(result_raw)
                if reagent_raw:
                    reagent_counts[reagent_raw] += 1
                    map_reagent_raw.setdefault(reagent_raw, {"raw_label": reagent_raw, "count": 0, "normalized_id": reagent_id, "mapping_status": reagent_status})
                    map_reagent_raw[reagent_raw]["count"] += 1
                if result_raw:
                    result_counts[result_raw] += 1
                observations.append({
                    "observation_id": stable_id("obs", test_id, ordinal),
                    "test_id": test_id,
                    "event_id": event["event_id"],
                    "source_sheet_name": event["source_sheet_name"],
                    "source_row": source_row,
                    "observation_ordinal": ordinal,
                    "substance_raw": raw["substance_raw"],
                    "substance_normalized_candidate": substance_id,
                    "reagent_raw": reagent_raw,
                    "reagent_normalized_candidate": reagent_id,
                    "reagent_map_status": reagent_status,
                    "result_raw": result_raw,
                    "result_normalized_candidate": result_candidate,
                    "result_map_status": result_status,
                    "observation_status": "source_observation_preserved",
                    "interpretation_policy": "contains_signal_only_no_claim_of_purity_or_dose",
                })

    source_sheets = []
    for event in events:
        source_sheets.append({
            "source_sheet_index": event["source_sheet_index"],
            "source_sheet_name": event["source_sheet_name"],
            "data_row_count_including_anomalies": len(sheet_data[event["source_sheet_name"]]) - 1,
            "source_sheet_hash": sheet_hash[event["source_sheet_name"]],
            "duplicate_group_id": event["duplicate_group_id"],
            "duplicate_group_size": event["duplicate_group_size"],
            "duplicate_status": event["duplicate_status"],
        })

    link_queue = []
    for event in events:
        for target_kind in ("venue", "producer"):
            link_queue.append({
                "link_id": stable_id("link", event["event_id"], target_kind),
                "event_id": event["event_id"],
                "source_sheet_name": event["source_sheet_name"],
                "target_kind": target_kind,
                "target_id": None,
                "target_name": None,
                "relation_type": "test_event_at_venue" if target_kind == "venue" else "test_event_by_producer",
                "evidence_ref": None,
                "confidence": "none",
                "status": "unlinked",
                "review_status": "pending_human_link",
                "not_inferred_from_sheet_name": True,
            })

    payload = {
        "schema_version": "rd-testing-event-evidence-v0.1",
        "generated_at": STAMP,
        "language": "en",
        "status": "candidate_evidence_pending_human_review",
        "source": {
            "file_name": SOURCE.name,
            "source_copy": SOURCE_COPY.name,
            "sha256": source_hash,
            "filename_period_label": source_period,
            "formula_count": 0,
        },
        "principles": [
            "Preserve source wording and source coordinates.",
            "Colorimetric observations are presence signals only; they do not establish identity, purity, dose, or safety.",
            "Event, venue, producer, and substance are separate identities.",
            "Venue and producer links remain unlinked until human evidence review.",
            "Exact duplicate source sheets are preserved and marked; they are not silently removed.",
        ],
        "source_sheets": source_sheets,
        "events": events,
        "test_rows": test_rows,
        "observations": observations,
        "substance_map": sorted(map_substance_raw.values(), key=lambda x: (-x["count"], x["raw_label"])),
        "reagent_map": sorted(map_reagent_raw.values(), key=lambda x: (-x["count"], x["raw_label"])),
        "link_queue": link_queue,
        "registries": {
            "venue_table": "future_venue_registry",
            "producer_table": "future_producer_registry",
            "relation_table": "link_queue",
        },
        "integration_contract": {
            "schema_version": "rd-testing-context-bridge-v0.1",
            "entity_keys": {
                "test": "test_id",
                "event": "event_id",
                "venue": "venue_id",
                "producer": "producer_id",
            },
            "future_tables": {
                "venue_registry": ["venue_id", "canonical_name", "aliases", "city", "source_refs", "status"],
                "producer_registry": ["producer_id", "canonical_name", "aliases", "source_refs", "status"],
                "event_venue_link": ["link_id", "event_id", "venue_id", "relation_type", "evidence_ref", "confidence", "status", "review_status"],
                "event_producer_link": ["link_id", "event_id", "producer_id", "relation_type", "evidence_ref", "confidence", "status", "review_status"],
            },
            "join_policy": [
                "Use stable IDs, never a display name as a foreign key.",
                "Keep event-to-venue and event-to-producer as separate many-to-many relations.",
                "Require an evidence_ref and human review before a link becomes public.",
                "Do not infer a venue or producer from a sheet name alone.",
                "Keep testing observations separate from public substance claims.",
            ],
        },
        "coverage": {
            "source_sheet_count": len(source_sheets),
            "event_count": len(events),
            "test_row_count_including_repeated_headers": len(test_rows),
            "test_row_count_non_header": sum(1 for x in test_rows if x["row_status"] != "repeated_header"),
            "test_row_count_data": sum(1 for x in test_rows if x["row_status"] == "data"),
            "observation_count": len(observations),
            "row_status_counts": dict(row_status_counts),
            "exact_duplicate_sheet_groups": sum(1 for x in groups.values() if len(x) > 1),
            "unresolved_substance_raw_labels": sum(1 for x in map_substance_raw.values() if x["mapping_status"] in {"unresolved_candidate", "misplaced_or_unresolved_candidate"}),
            "unresolved_reagent_raw_labels": sum(1 for x in map_reagent_raw.values() if x["mapping_status"] == "unresolved_candidate"),
        },
        "review_queue": [
            "Confirm which source sheets belong to the 2025 period when the filename and sheet labels conflict.",
            "Review exact duplicate sheets before public aggregation; retain raw evidence.",
            "Resolve substance and reagent candidates marked unresolved or misplaced.",
            "Link event_id to venue_id and producer_id only with explicit evidence.",
            "Have an RD human gate approve any public interpretation of color changes.",
        ],
    }
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")

    build_workbook(payload)
    write_report(payload)
    print(json.dumps({"json": str(JSON_OUT), "xlsx": str(XLSX_OUT), "report": str(REPORT_OUT), "source_copy": str(SOURCE_COPY), "coverage": payload["coverage"]}, ensure_ascii=True, indent=2))


def build_workbook(payload: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    def add_sheet(name: str, rows: list[dict[str, Any]], columns: list[str], comments: dict[str, str] | None = None) -> None:
        ws = wb.create_sheet(name)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for col, key in enumerate(columns, start=1):
            cell = ws.cell(1, col, key)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if comments and key in comments:
                cell.comment = Comment(comments[key], "RD integration")
        for rindex, record in enumerate(rows, start=2):
            for cindex, key in enumerate(columns, start=1):
                cell = ws.cell(rindex, cindex, record.get(key))
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
        for cindex, key in enumerate(columns, start=1):
            max_len = len(key)
            for row in rows[:500]:
                value = row.get(key)
                if value is not None:
                    max_len = max(max_len, min(60, len(str(value))))
            ws.column_dimensions[get_column_letter(cindex)].width = min(60, max(12, max_len + 2))

    manifest = [
        {"key": "schema_version", "value": payload["schema_version"], "note": "Machine-readable evidence schema."},
        {"key": "source_file", "value": payload["source"]["file_name"], "note": f"SHA256: {payload['source']['sha256']}"},
        {"key": "source_copy", "value": payload["source"]["source_copy"], "note": "Original copied without editing."},
        {"key": "interpretation_rule", "value": "contains_signal_only", "note": "A color test is not identity, purity, dose, or safety."},
        {"key": "event_link_rule", "value": "human_evidence_required", "note": "Do not infer venue or producer from a sheet name."},
        {"key": "duplicate_rule", "value": "preserve_and_mark", "note": "Exact duplicate sheets remain available as source evidence."},
        {"key": "period_rule", "value": "filename_2025_but_sheet_dates_reviewed", "note": "Some sheet labels contain 2026 candidates and remain flagged."},
    ]
    add_sheet("Manifest", manifest, ["key", "value", "note"])
    add_sheet("Source Sheets", payload["source_sheets"], list(payload["source_sheets"][0].keys()))
    add_sheet("Events", payload["events"], list(payload["events"][0].keys()))
    add_sheet("Test Rows", payload["test_rows"], list(payload["test_rows"][0].keys()))
    add_sheet("Observations", payload["observations"], list(payload["observations"][0].keys()))
    add_sheet("Substance Map", payload["substance_map"], list(payload["substance_map"][0].keys()))
    add_sheet("Reagent Map", payload["reagent_map"], list(payload["reagent_map"][0].keys()))
    add_sheet("Link Queue", payload["link_queue"], list(payload["link_queue"][0].keys()))
    quality = [{"metric": k, "value": v, "interpretation": "candidate_count_from_source_preserved"} for k, v in payload["coverage"].items() if not isinstance(v, dict)]
    quality.extend({"metric": f"row_status:{k}", "value": v, "interpretation": "classification_requires_review"} for k, v in payload["coverage"]["row_status_counts"].items())
    add_sheet("Data Quality", quality, ["metric", "value", "interpretation"])
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.save(XLSX_OUT)


def write_report(payload: dict[str, Any]) -> None:
    coverage = payload["coverage"]
    rows = [
        f"# Integración de testeos en eventos — base 2025",
        "",
        f"Generado: {STAMP}",
        "",
        "## Resultado",
        "",
        f"Se preservó la fuente `Testeo 2025.xlsx` y se construyó una integración separada con {coverage['source_sheet_count']} pestañas, {coverage['event_count']} eventos-fuente, {coverage['test_row_count_non_header']} filas no vacías de fuente (de ellas, {coverage['test_row_count_data']} quedan clasificadas como datos y {coverage['row_status_counts'].get('data_with_unresolved_substance', 0)} como datos con anomalías) y {coverage['observation_count']} observaciones de reactivo/resultado.",
        "",
        "La fuente no se interpretó como una tabla de composición. Cada registro conserva su pestaña y fila original; las normalizaciones son candidatas para revisión humana.",
        "",
        "## Regla sanitaria y semántica",
        "",
        "> Los testeos colorimétricos registran una señal de presencia compatible con ciertos compuestos o familias. No demuestran identidad definitiva, pureza, potencia, cantidad ni seguridad.",
        "",
        "Por eso la integración usa `contains_signal_only` y conserva el texto original de cada resultado.",
        "",
        "## Puente futuro",
        "",
        "La conexión queda preparada sin inventar vínculos:",
        "",
        "`test_id -> event_id -> venue_id / producer_id`",
        "",
        "- `event_id` identifica la pestaña-fuente, no una venue ni una productora.",
        "- `venue_id` y `producer_id` quedan vacíos hasta contar con evidencia explícita.",
        "- `Link Queue` contiene dos enlaces pendientes por evento: evento-venue y evento-productora.",
        "- Se conserva `evidence_ref`, `confidence`, `status` y `review_status` para una futura vinculación humana.",
        "",
        "## Hallazgos de calidad",
        "",
        f"- Filas de datos incluyendo anomalías: {coverage['test_row_count_including_repeated_headers']}",
        f"- Filas no vacías sin encabezados repetidos: {coverage['test_row_count_non_header']}",
        f"- Filas clasificadas como datos: {coverage['test_row_count_data']}",
        f"- Grupos de pestañas con contenido idéntico: {coverage['exact_duplicate_sheet_groups']}",
        f"- Filas con sustancia ausente o no resuelta: {coverage['row_status_counts'].get('data_with_unresolved_substance', 0)}",
        f"- Etiquetas de sustancia no resueltas: {coverage['unresolved_substance_raw_labels']}",
        f"- Etiquetas de reactivo no resueltas: {coverage['unresolved_reagent_raw_labels']}",
        "",
        "El nombre del archivo dice 2025, pero algunas pestañas contienen tokens de fecha que parecen corresponder a 2026. No se eliminaron: quedaron marcadas en `Events` para confirmar el periodo antes de publicar agregados.",
        "",
        "## Archivos",
        "",
        f"- Fuente preservada: `{SOURCE_COPY.name}`",
        f"- Evidencia machine-readable: `{JSON_OUT.name}`",
        f"- Libro integrado: `{XLSX_OUT.name}`",
        f"- Esta síntesis: `{REPORT_OUT.name}`",
        "",
        "## Próxima revisión humana",
        "",
        "1. Confirmar qué pestañas pertenecen efectivamente a 2025.",
        "2. Revisar duplicados exactos y decidir si se excluyen solo de agregados públicos.",
        "3. Resolver sustancias y reactivos marcados como candidatos o no resueltos.",
        "4. Vincular cada evento con venue/productora usando evidencia, sin inferir desde el nombre de la pestaña.",
        "5. Aprobar el lenguaje público de cada resultado antes de llevarlo a la matriz o a un POST.",
        "",
    ]
    REPORT_OUT.write_text("\n".join(rows), encoding="utf-8")


if __name__ == "__main__":
    main()
